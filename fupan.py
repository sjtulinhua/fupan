import os
import sys
import shutil
import argparse
import datetime
import pandas as pd
import configs as cfg
import colordic as cdic
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, colors


g_dest_col_num = 1
g_database_df_dict = {}


def open_dest_xlsx(fpath):
    # open xlsx
    return load_workbook(fpath)


def backup_before_proc(orig_file, bk_file):
    shutil.copyfile(orig_file, bk_file)


def get_dest_col():
    return g_dest_col_num


def insert_data_by_col(workbook, sheet_num, col_idx, color_map, data):
    ws = workbook.worksheets[sheet_num]
    ws.insert_cols(col_idx)

    # setup cell format
    for idx, val in enumerate(data, start=0):
        # print(col_idx, idx, val)
        c = ws.cell(column=col_idx, row=idx + 1, value=val)

        if idx == 0:
            c.fill = PatternFill(fill_type='solid', fgColor='d1ffbd')
        elif val in color_map.keys():
            c.fill = PatternFill(fill_type='solid', fgColor=color_map[val])
        else:
            c.fill = PatternFill(fill_type='solid', fgColor='fff39a')

        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.font = Font(name='宋体', size=16)
        c.border = Border(left=Side(style='thin', color=colors.BLACK),
                          right=Side(style='thin', color=colors.BLACK),
                          top=Side(style='thin', color=colors.BLACK),
                          bottom=Side(style='thin', color=colors.BLACK))
        # setup row height
        ws.row_dimensions[idx + 1].height = 25


def insert_data(workbook, sheet_num, col, col_data_list, color_map):
    """
    :type col_data_list: list
    """
    for data in col_data_list:
        insert_data_by_col(workbook, sheet_num, col, color_map, data)


def data_grouping(data_df, date, database):
    short_date_str = date.strftime('%m%d')

    if database == 'tdx':
        cnt_df = data_df.groupby('细分行业')['代码'].count().to_frame(name='count')
        cnt_df.sort_values(['count'], ascending=False, inplace=True)

        # add current date to dataframe
        cnt_df.rename(columns={'count': short_date_str}, inplace=True)
        cnt_df.reset_index(inplace=True)

        # prepare data for writing to the xlsx
        col_cnt = pd.Series(cnt_df.columns.array[1])
        col_cnt = col_cnt.append(cnt_df[short_date_str])
        col_industry = pd.Series(cnt_df.columns.array[0])
        col_industry = col_industry.append(cnt_df['细分行业'])

        # percentage
        size = []
        for industry in cnt_df['细分行业']:
            size.append(cfg.board_size_tdx[industry])
        cnt_df['board size'] = size
        cnt_df['占比'] = cnt_df[short_date_str] / cnt_df['board size']
        cnt_df['占比'] = cnt_df['占比'].apply(lambda x: format(x, '.1%'))
        col_pct = pd.Series(cnt_df.columns.array[3])
        col_pct = col_pct.append(cnt_df['占比'])
        return col_cnt, col_industry, col_pct

    elif database == 'futu':
        # todo remove from the loops
        # sql_cmd = "select * from concept_dashboard_name"
        # context = sqlite3.connect(cfg.database_futu)
        # futu_db_df = pd.read_sql(sql_cmd, context, index_col='index')
        # for concept in cfg.exclude_concepts_futu:
        #     futu_db_df.drop(concept, axis=1, inplace=True)
        global g_database_df_dict
        futu_db_df = g_database_df_dict[database]

        # clean data (looks no need for db=tdx where these dirty data will be cleaned by groupby
        data_df.drop(data_df.index[-1], axis=0, inplace=True)
        data_df.drop([data_df.columns[-1]], axis=1, inplace=True)
        stock_list_df = data_df[['名称']]
        stock_concept_dashboard = pd.DataFrame()

        for idx, row in stock_list_df.iterrows():
            concept_dict = lookup_concept(row['名称'], futu_db_df)
            stock_concept_row_df = pd.DataFrame.from_dict(concept_dict)
            stock_concept_row_df.insert(0, '名称', row['名称'])
            stock_concept_dashboard = stock_concept_dashboard.append(stock_concept_row_df, ignore_index=True,
                                                                     sort=False)
        sum_df = pd.DataFrame()
        for label, content in stock_concept_dashboard.iteritems():
            # print(label, sep='\n')
            if label == '名称':
                continue
            else:
                group_col_df = stock_concept_dashboard.groupby(label)['名称'].count().to_frame(name='count')
                sum_df = sum_df.add(group_col_df, fill_value=0)
        sum_df.sort_values(['count'], ascending=False, inplace=True)
        sum_df.reset_index(inplace=True)
        sum_df.rename(columns={'index': 'Futu概念', 'count': short_date_str}, inplace=True)

        # prepare data for writing to the xlsx
        col_cnt = pd.Series(sum_df.columns.array[1])
        col_cnt = col_cnt.append(sum_df[short_date_str])
        col_industry = pd.Series(sum_df.columns.array[0])
        col_industry = col_industry.append(sum_df['Futu概念'])

        # percentage
        plate_size_dict = futu_db_df.count().to_dict()
        size = []
        for concept in sum_df['Futu概念']:
            size.append(plate_size_dict[concept])
        sum_df['board size'] = size
        sum_df['占比'] = sum_df[short_date_str] / sum_df['board size']
        sum_df['占比'] = sum_df['占比'].apply(lambda x: format(x, '.1%'))

        col_pct = pd.Series(sum_df.columns.array[3])
        col_pct = col_pct.append(sum_df['占比'])

        return col_cnt, col_industry, col_pct


def update_analysis(datafpath, workbook, sheetnum, date, database):
    if not os.path.exists(datafpath):
        print(f"Skip non-existing file: {datafpath} ...")
        return

    # start zt analysis
    print(f"processing {datafpath}")

    # read data
    df = pd.read_csv(datafpath, sep='\t', encoding="gbk")

    # do analysis
    col_cnt, col_industry, col_pct = data_grouping(df, date, database)

    if database == 'tdx':
        color_map = cdic.Paired_color_map
    elif database == 'futu':
        color_map = cdic.Paired400_color_map
    else:
        color_map = cdic.Paired_color_map

    insert_data(workbook, sheetnum, get_dest_col(), [col_pct, col_cnt, col_industry], color_map) # shall use the right col order


    # setup col width
    # todo iterate all columns then setup width separately
    # ws.column_dimensions['A'].width = 12
    # ws.column_dimensions['B'].width = 6.5


def lookup_concept(stock_name, plate_dashboard_df):
    # todo the most time-consuming function, need to be optimized
    concept_dict = {}
    index = 1
    for label, content in plate_dashboard_df.iteritems():
        if stock_name in content.values:
            concept_dict.update({'concept'+str(index): [label]})
            index += 1
    return concept_dict

def gen_report(args, workbook, database):
    sdate_input = datetime.datetime.strptime(args.sdate, '%Y%m%d')
    if args.edate is not None:
        edate_input = datetime.datetime.strptime(args.edate, '%Y%m%d')
        for i in range((edate_input - sdate_input).days + 1):
            day = sdate_input + datetime.timedelta(days=i)
            str_day = day.strftime('%Y%m%d')
            update_analysis(cfg.sheet_zt + str_day + '.txt', workbook, cfg.sheet_zt_num, day, database)
            update_analysis(cfg.sheet_lsxg + str_day + '.txt', workbook, cfg.sheet_lsxg_num, day, database)
            update_analysis(cfg.sheet_drps + str_day + '.txt', workbook, cfg.sheet_drps_num, day, database)
            update_analysis(cfg.sheet_srps + str_day + '.txt', workbook, cfg.sheet_srps_num, day, database)
    else:
        str_day = args.sdate
        update_analysis(cfg.sheet_zt + str_day + '.txt', workbook, cfg.sheet_zt_num, sdate_input, database)
        update_analysis(cfg.sheet_lsxg + str_day + '.txt', workbook, cfg.sheet_lsxg_num, sdate_input, database)
        update_analysis(cfg.sheet_drps + str_day + '.txt', workbook, cfg.sheet_drps_num, sdate_input, database)
        update_analysis(cfg.sheet_srps + str_day + '.txt', workbook, cfg.sheet_srps_num, sdate_input, database)

    return workbook


def load_database(database):
    global g_database_df_dict
    if database == 'futu':
        sql_cmd = "select * from concept_dashboard_name"
        context = sqlite3.connect(cfg.database_futu)
        futu_db_df = pd.read_sql(sql_cmd, context, index_col='index')
        for concept in cfg.exclude_concepts_futu:
            futu_db_df.drop(concept, axis=1, inplace=True)
        g_database_df_dict.update({database: futu_db_df})

    elif database == 'tushare':
        print('not implemented database for tushare yet')

def fupan_main(database):
    parser = argparse.ArgumentParser(description='命令行中传入"YYYYMMDD"格式的日期')
    parser.add_argument("sdate", help="input the start date in the format of 'YYYYMMDD'")
    parser.add_argument("--edate", required=False, help="input the end date in the format of 'YYYYMMDD'")
    parser.add_argument("--destcol", type=int, required=False, help="input the num of col (only accept number) "
                                                                    "before which you want to insert new columns")
    args = parser.parse_args()

    if args.destcol is not None:
        print("optional arg: destcol = " + str(args.destcol))
        global g_dest_col_num
        g_dest_col_num = args.destcol


    for db in database:
        if db == 'tdx':
            print(f'using database:  {db}')
            backup_before_proc(cfg.dest_xlsx_tdx, cfg.backup_xlsx_tdx)
            workbook = open_dest_xlsx(cfg.dest_xlsx_tdx)
            gen_report(args, workbook, db)
            print(f"Completed, write to {cfg.dest_xlsx_tdx}!")
            workbook.save(cfg.dest_xlsx_tdx)

        elif db == 'futu':
            print(f'using database:  {db}')
            backup_before_proc(cfg.dest_xlsx_futu, cfg.backup_xlsx_futu)
            load_database(db)
            workbook = open_dest_xlsx(cfg.dest_xlsx_futu)
            gen_report(args, workbook, db)
            print(f"Completed, write to {cfg.dest_xlsx_futu}!")
            workbook.save(cfg.dest_xlsx_futu)

        elif db == 'tushare':
            # todo
            print(f'using database:  {db}')


if __name__ == '__main__':
    fupan_main(cfg.database)
